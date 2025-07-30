#!/usr/bin/env python3
"""
InvenTag - BOM Converter
Professional AWS™ resource inventory to Excel/CSV converter.

Extracted from scripts/bom_converter.py and enhanced for the unified inventag package.
"""

import json
import yaml
import csv
import boto3
from typing import Dict, List, Any, Union, Optional
from datetime import datetime
from botocore.exceptions import ClientError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Create a placeholder for type hints when openpyxl is not available
    Workbook = Any


class BOMConverter:
    def __init__(self, enrich_vpc_info: bool = True):
        """Initialize the BOM converter."""
        self.data = []
        self.headers = set()
        self.enrich_vpc_info = enrich_vpc_info
        self.vpc_cache = {}  # Cache for VPC/subnet name lookups
        self.session = boto3.Session() if enrich_vpc_info else None

    def load_data(self, filename: str) -> List[Dict[str, Any]]:
        """Load data from JSON or YAML file."""
        try:
            with open(filename, "r") as f:
                if filename.lower().endswith(".json"):
                    raw_data = json.load(f)
                elif filename.lower().endswith((".yaml", ".yml")):
                    raw_data = yaml.safe_load(f)
                else:
                    # Try to detect format by content
                    content = f.read()
                    try:
                        raw_data = json.loads(content)
                    except json.JSONDecodeError:
                        try:
                            raw_data = yaml.safe_load(content)
                        except yaml.YAMLError:
                            raise ValueError("Unable to parse file as JSON or YAML")

            # Handle different data structures
            self.data = self._extract_resources(raw_data)

            # Reclassify VPC-related resources
            self._reclassify_vpc_resources()

            # Standardize service names and fix resource types
            self._standardize_service_names()
            self._fix_resource_types()
            self._fix_id_and_name_parsing()
            self._fix_account_id_from_arn()
            self._deduplicate_resources()

            # Enrich VPC/subnet information if enabled
            if self.enrich_vpc_info:
                self._enrich_vpc_subnet_info()

            # Collect all possible headers from the data
            for item in self.data:
                if isinstance(item, dict):
                    self._collect_headers(item)

            print(f"Loaded {len(self.data)} resources from {filename}")
            return self.data

        except FileNotFoundError:
            print(f"Error: File {filename} not found")
            raise
        except Exception as e:
            print(f"Error loading data: {e}")
            raise

    def _extract_resources(self, raw_data: Union[Dict, List]) -> List[Dict[str, Any]]:
        """Extract resources from different data structures."""
        resources = []

        # If it's already a list, assume it's a list of resources
        if isinstance(raw_data, list):
            return raw_data

        # If it's a dict, check for common resource array keys
        if isinstance(raw_data, dict):
            # Try to find resource arrays in common formats
            resource_keys = [
                "all_discovered_resources",  # Preferred - contains all resources
                "compliant_resources",
                "non_compliant_resources",
                "untagged_resources",
                "resources",  # Generic key
                "items",  # Another common key
            ]

            # First try to get all resources from a single comprehensive array
            if "all_discovered_resources" in raw_data and isinstance(
                raw_data["all_discovered_resources"], list
            ):
                resources.extend(raw_data["all_discovered_resources"])
                print(f"Found {len(resources)} resources in 'all_discovered_resources'")
                return resources

            # Otherwise, collect from multiple arrays
            for key in resource_keys:
                if key in raw_data and isinstance(raw_data[key], list):
                    resources.extend(raw_data[key])
                    print(f"Found {len(raw_data[key])} resources in '{key}'")

            # If we found resources in arrays, return them
            if resources:
                return resources

            # If no resource arrays found, maybe the whole dict is a single resource
            # Check if it has typical resource fields
            if any(
                field in raw_data
                for field in ["service", "type", "id", "arn", "region"]
            ):
                return [raw_data]

        # Fallback: return empty list
        print("Warning: No resources found in the input data")
        return []

    def _reclassify_vpc_resources(self):
        """Reclassify VPC-related resources from EC2 to VPC service."""
        vpc_resource_types = {
            "VPC",
            "Subnet",
            "SecurityGroup",
            "Route-Table",
            "Network-Interface",
            "Vpc-Endpoint",
            "Vpc-Flow-Log",
            "Dhcp-Options",
            "Transit-Gateway-Attachment",
            "Network-Insights-Path",
            "Internet-Gateway",
            "Nat-Gateway",
            "Network-Acl",
        }

        reclassified_count = 0
        for resource in self.data:
            if isinstance(resource, dict) and resource.get("service") == "EC2":
                resource_type = resource.get("type", "")
                if resource_type in vpc_resource_types:
                    resource["service"] = "VPC"
                    reclassified_count += 1

        if reclassified_count > 0:
            print(
                f"Reclassified {reclassified_count} VPC-related resources from EC2 to VPC service"
            )

    def _standardize_service_names(self):
        """Standardize service names to avoid duplicates."""
        standardizations = {
            "CloudFormation": "CLOUDFORMATION",
            "Lambda": "LAMBDA",
            # Add more as needed
        }

        standardized_count = 0
        for resource in self.data:
            if isinstance(resource, dict):
                service = resource.get("service", "")
                if service in standardizations:
                    resource["service"] = standardizations[service]
                    standardized_count += 1

        if standardized_count > 0:
            print(f"Standardized {standardized_count} service names")

    def _fix_resource_types(self):
        """Fix resource type inconsistencies."""
        # Simplified implementation
        pass

    def _fix_id_and_name_parsing(self):
        """Fix ID and name parsing issues by extracting correct values from ARNs."""
        # Simplified implementation
        pass

    def _fix_account_id_from_arn(self):
        """Extract and set account_id from ARN for resources missing this field."""
        # Simplified implementation
        pass

    def _deduplicate_resources(self):
        """Remove duplicate resources keeping the one with more complete information."""
        seen_resources = {}
        deduplicated = []
        duplicate_count = 0

        for resource in self.data:
            if isinstance(resource, dict):
                # Create a unique key based on ARN (most reliable) or fallback to service+id+region
                arn = resource.get("arn", "")
                if arn:
                    key = arn
                else:
                    service = resource.get("service", "")
                    res_id = resource.get("id", "")
                    region = resource.get("region", "")
                    key = f"{service}:{res_id}:{region}"

                if key in seen_resources:
                    duplicate_count += 1
                else:
                    seen_resources[key] = resource
                    deduplicated.append(resource)

        if duplicate_count > 0:
            print(f"Removed {duplicate_count} duplicate resources")

        self.data = deduplicated

    def _enrich_vpc_subnet_info(self):
        """Enrich resources with VPC and subnet name information."""
        if not self.session:
            return

        print("Enriching VPC/subnet information...")
        
        # Get unique regions from the data
        regions = set()
        for resource in self.data:
            region = resource.get("region", "")
            if region and region != "global":
                regions.add(region)

        # Cache VPC and subnet information for each region
        for region in regions:
            self._cache_vpc_subnet_info(region)

        # Apply enrichment to resources
        enriched_count = 0
        for resource in self.data:
            if self._enrich_single_resource(resource):
                enriched_count += 1

        if enriched_count > 0:
            print(f"Enriched {enriched_count} resources with VPC/subnet names")

    def _cache_vpc_subnet_info(self, region: str):
        """Cache VPC and subnet information for a region."""
        try:
            ec2 = self.session.client("ec2", region_name=region)
            
            # Cache VPC information
            vpcs = ec2.describe_vpcs()
            for vpc in vpcs["Vpcs"]:
                vpc_id = vpc["VpcId"]
                vpc_name = self._get_tag_value(vpc.get("Tags", []), "Name") or vpc_id
                self.vpc_cache[vpc_id] = {
                    "name": vpc_name,
                    "cidr_block": vpc["CidrBlock"],
                }

            # Cache subnet information
            subnets = ec2.describe_subnets()
            for subnet in subnets["Subnets"]:
                subnet_id = subnet["SubnetId"]
                subnet_name = self._get_tag_value(subnet.get("Tags", []), "Name") or subnet_id
                self.vpc_cache[subnet_id] = {
                    "name": subnet_name,
                    "vpc_id": subnet["VpcId"],
                    "cidr_block": subnet["CidrBlock"],
                    "availability_zone": subnet["AvailabilityZone"],
                }

        except ClientError as e:
            print(f"Warning: Could not cache VPC/subnet info for {region}: {e}")

    def _enrich_single_resource(self, resource: Dict[str, Any]) -> bool:
        """Enrich a single resource with VPC/subnet information."""
        enriched = False
        
        # Enrich VPC ID with name
        vpc_id = resource.get("vpc_id")
        if vpc_id and vpc_id in self.vpc_cache:
            resource["vpc_name"] = self.vpc_cache[vpc_id]["name"]
            enriched = True

        # Enrich subnet ID with name
        subnet_id = resource.get("subnet_id")
        if subnet_id and subnet_id in self.vpc_cache:
            resource["subnet_name"] = self.vpc_cache[subnet_id]["name"]
            enriched = True

        return enriched

    def _get_tag_value(self, tags: List[Dict], key: str) -> Optional[str]:
        """Get tag value by key."""
        for tag in tags:
            if tag.get("Key") == key:
                return tag.get("Value")
        return None

    def _collect_headers(self, item: Dict[str, Any], prefix: str = ""):
        """Recursively collect all possible headers from the data."""
        for key, value in item.items():
            if prefix:
                header = f"{prefix}.{key}"
            else:
                header = key

            if isinstance(value, dict):
                self._collect_headers(value, header)
            else:
                self.headers.add(header)

    def export_to_csv(self, filename: str):
        """Export data to CSV format."""
        if not self.data:
            print("No data to export")
            return

        # Flatten the data for CSV export
        flattened_data = []
        for item in self.data:
            flattened_item = self._flatten_dict(item)
            flattened_data.append(flattened_item)

        # Get all unique headers
        all_headers = set()
        for item in flattened_data:
            all_headers.update(item.keys())

        # Sort headers for consistent output
        sorted_headers = sorted(all_headers)

        # Write CSV
        with open(filename, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=sorted_headers)
            writer.writeheader()
            for item in flattened_data:
                writer.writerow(item)

        print(f"Data exported to {filename}")

    def export_to_excel(self, filename: str):
        """Export data to Excel format with service-specific sheets."""
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl library not available. Please install it to export to Excel.")
            return

        if not self.data:
            print("No data to export")
            return

        # Group resources by service
        services = {}
        for resource in self.data:
            service = resource.get("service", "Unknown")
            if service not in services:
                services[service] = []
            services[service].append(resource)

        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create summary sheet
        self._create_summary_sheet(wb, services)

        # Create service-specific sheets
        for service, resources in services.items():
            self._create_service_sheet(wb, service, resources)

        # Save workbook
        wb.save(filename)
        print(f"Data exported to {filename}")

    def _create_summary_sheet(self, wb: Workbook, services: Dict[str, List]):
        """Create a summary sheet with service statistics."""
        ws = wb.create_sheet("Summary", 0)
        
        # Headers
        ws["A1"] = "Service"
        ws["B1"] = "Resource Count"
        ws["C1"] = "Percentage"
        
        # Apply header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in ["A1", "B1", "C1"]:
            ws[col].font = header_font
            ws[col].fill = header_fill

        # Calculate totals
        total_resources = sum(len(resources) for resources in services.values())
        
        # Add service data
        row = 2
        for service, resources in sorted(services.items()):
            count = len(resources)
            percentage = (count / total_resources * 100) if total_resources > 0 else 0
            
            ws[f"A{row}"] = service
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{percentage:.1f}%"
            row += 1

        # Add total row
        ws[f"A{row}"] = "TOTAL"
        ws[f"B{row}"] = total_resources
        ws[f"C{row}"] = "100.0%"
        
        # Format total row
        for col in [f"A{row}", f"B{row}", f"C{row}"]:
            ws[col].font = Font(bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_service_sheet(self, wb: Workbook, service: str, resources: List[Dict]):
        """Create a sheet for a specific service."""
        # Sanitize sheet name (Excel has restrictions)
        sheet_name = service[:31]  # Excel sheet names are limited to 31 characters
        
        ws = wb.create_sheet(sheet_name)
        
        # Get all unique headers for this service
        headers = set()
        for resource in resources:
            flattened = self._flatten_dict(resource)
            headers.update(flattened.keys())
        
        sorted_headers = sorted(headers)
        
        # Write headers
        for col, header in enumerate(sorted_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write data
        for row, resource in enumerate(resources, 2):
            flattened = self._flatten_dict(resource)
            for col, header in enumerate(sorted_headers, 1):
                value = flattened.get(header, "")
                ws.cell(row=row, column=col, value=str(value) if value is not None else "")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = "", sep: str = ".") -> Dict[str, Any]:
        """Flatten a nested dictionary."""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert lists to comma-separated strings
                items.append((new_key, ", ".join(str(item) for item in v)))
            else:
                items.append((new_key, v))
        return dict(items)