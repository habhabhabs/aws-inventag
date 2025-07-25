#!/usr/bin/env python3
"""
BOM Converter Utility
Converts JSON/YAML AWS resource inventory to Excel/CSV format.
Uses minimal dependencies for portability.
"""

import argparse
import json
import yaml
import csv
import sys
from typing import Dict, List, Any, Union
from datetime import datetime
import os


class BOMConverter:
    def __init__(self):
        """Initialize the BOM converter."""
        self.data = []
        self.headers = set()
    
    def load_data(self, filename: str) -> List[Dict[str, Any]]:
        """Load data from JSON or YAML file."""
        try:
            with open(filename, 'r') as f:
                if filename.lower().endswith('.json'):
                    self.data = json.load(f)
                elif filename.lower().endswith(('.yaml', '.yml')):
                    self.data = yaml.safe_load(f)
                else:
                    # Try to detect format by content
                    content = f.read()
                    try:
                        self.data = json.loads(content)
                    except json.JSONDecodeError:
                        try:
                            self.data = yaml.safe_load(content)
                        except yaml.YAMLError:
                            raise ValueError("Unable to parse file as JSON or YAML")
            
            # Collect all possible headers from the data
            for item in self.data:
                if isinstance(item, dict):
                    self._collect_headers(item)
            
            print(f"Loaded {len(self.data)} resources from {filename}")
            return self.data
            
        except FileNotFoundError:
            print(f"Error: File {filename} not found")
            sys.exit(1)
        except Exception as e:
            print(f"Error loading data: {e}")
            sys.exit(1)
    
    def _collect_headers(self, item: Dict[str, Any], prefix: str = ""):
        """Recursively collect all possible headers from nested dictionaries."""
        for key, value in item.items():
            header = f"{prefix}{key}" if prefix else key
            
            if isinstance(value, dict):
                # For nested dictionaries, create flattened headers
                self._collect_headers(value, f"{header}.")
            elif isinstance(value, list):
                # For lists, just add the header (we'll stringify the list)
                self.headers.add(header)
            else:
                self.headers.add(header)
    
    def _flatten_dict(self, item: Dict[str, Any], prefix: str = "") -> Dict[str, str]:
        """Flatten nested dictionaries for tabular output."""
        flattened = {}
        
        for key, value in item.items():
            header = f"{prefix}{key}" if prefix else key
            
            if isinstance(value, dict):
                flattened.update(self._flatten_dict(value, f"{header}."))
            elif isinstance(value, list):
                # Convert lists to comma-separated strings
                flattened[header] = ", ".join(str(v) for v in value)
            elif value is None:
                flattened[header] = ""
            else:
                flattened[header] = str(value)
        
        return flattened
    
    def to_csv(self, output_filename: str):
        """Convert data to CSV format."""
        try:
            # Sort headers for consistent output
            sorted_headers = sorted(list(self.headers))
            
            with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=sorted_headers)
                writer.writeheader()
                
                for item in self.data:
                    if isinstance(item, dict):
                        flattened = self._flatten_dict(item)
                        # Ensure all headers are present (fill missing with empty string)
                        row = {header: flattened.get(header, '') for header in sorted_headers}
                        writer.writerow(row)
            
            print(f"CSV export saved to {output_filename}")
            
        except Exception as e:
            print(f"Error creating CSV: {e}")
            sys.exit(1)
    
    def to_excel(self, output_filename: str):
        """Convert data to Excel format using openpyxl."""
        try:
            # Try to import openpyxl, fall back to CSV if not available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter
            except ImportError:
                print("Warning: openpyxl not available. Falling back to CSV format.")
                csv_filename = output_filename.replace('.xlsx', '.csv')
                self.to_csv(csv_filename)
                return
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "AWS Resources"
            
            # Sort headers for consistent output
            sorted_headers = sorted(list(self.headers))
            
            # Write headers with formatting
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(sorted_headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            
            # Write data rows
            for row_num, item in enumerate(self.data, 2):
                if isinstance(item, dict):
                    flattened = self._flatten_dict(item)
                    for col_num, header in enumerate(sorted_headers, 1):
                        value = flattened.get(header, '')
                        ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for col_num, header in enumerate(sorted_headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(header)
                
                # Check data in this column to find max length
                for row_num in range(2, min(102, len(self.data) + 2)):  # Check first 100 rows
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # Set column width (max 50 to avoid extremely wide columns)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create summary worksheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws.append(["AWS Resource Inventory Summary"])
            summary_ws.append([])
            summary_ws.append(["Generated on:", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')])
            summary_ws.append(["Total Resources:", len(self.data)])
            summary_ws.append([])
            
            # Resource count by service
            service_counts = {}
            region_counts = {}
            type_counts = {}
            
            for item in self.data:
                if isinstance(item, dict):
                    service = item.get('service', 'Unknown')
                    region = item.get('region', 'Unknown')
                    resource_type = item.get('type', 'Unknown')
                    
                    service_counts[service] = service_counts.get(service, 0) + 1
                    region_counts[region] = region_counts.get(region, 0) + 1
                    type_counts[resource_type] = type_counts.get(resource_type, 0) + 1
            
            summary_ws.append(["Resources by Service:"])
            for service, count in sorted(service_counts.items()):
                summary_ws.append([service, count])
            
            summary_ws.append([])
            summary_ws.append(["Resources by Region:"])
            for region, count in sorted(region_counts.items()):
                summary_ws.append([region, count])
            
            summary_ws.append([])
            summary_ws.append(["Resources by Type:"])
            for resource_type, count in sorted(type_counts.items()):
                summary_ws.append([resource_type, count])
            
            # Format summary sheet
            summary_ws.cell(1, 1).font = Font(bold=True, size=14)
            
            # Save workbook
            wb.save(output_filename)
            print(f"Excel export saved to {output_filename}")
            
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description='BOM Converter - Convert JSON/YAML to Excel/CSV')
    parser.add_argument('--input', '-i', required=True, 
                       help='Input JSON or YAML file')
    parser.add_argument('--output', '-o', required=True, 
                       help='Output filename')
    parser.add_argument('--format', choices=['excel', 'csv'], default='excel',
                       help='Output format (default: excel)')
    
    args = parser.parse_args()
    
    # Validate input file exists
    if not os.path.exists(args.input):
        print(f"Error: Input file {args.input} does not exist")
        sys.exit(1)
    
    # Initialize converter
    converter = BOMConverter()
    
    # Load data
    converter.load_data(args.input)
    
    # Convert based on format
    if args.format == 'excel':
        # Ensure output has .xlsx extension
        if not args.output.lower().endswith('.xlsx'):
            args.output += '.xlsx'
        converter.to_excel(args.output)
    elif args.format == 'csv':
        # Ensure output has .csv extension
        if not args.output.lower().endswith('.csv'):
            args.output += '.csv'
        converter.to_csv(args.output)


if __name__ == "__main__":
    main()