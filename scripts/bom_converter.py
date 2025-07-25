#!/usr/bin/env python3
"""
AWS InvenTag - BOM Converter
Professional AWS resource inventory to Excel/CSV converter.

Part of AWS InvenTag: Python tool to check on AWS cloud inventory and tagging.
Integrate into your CI/CD flow to meet your organization's stringent compliance requirements.

Features:
- Creates service-specific Excel sheets for organized reporting
- Intelligent data enhancement with VPC names and account IDs
- Standardizes resource types and IDs across discovery methods
- Professional formatting with summary dashboards
- CI/CD ready for automated compliance workflows

Usage:
    python bom_converter.py --input inventory.json --output report.xlsx
"""

import argparse
import json
import yaml
import csv
import sys
import boto3
from typing import Dict, List, Any, Union, Optional
from datetime import datetime
import os
from botocore.exceptions import ClientError


class BOMConverter:
    def __init__(self, enrich_vpc_info: bool = True):
        """Initialize the BOM converter."""
        self.data = []
        self.headers = set()
        self.enrich_vpc_info = enrich_vpc_info
        self.vpc_cache = {}  # Cache for VPC/subnet name lookups
        self.session = boto3.Session() if enrich_vpc_info else None
    
    def load_data(self, filename: str) -> List[Dict[str, Any]]:
        """Load data from JSON or YAML file."""
        try:
            with open(filename, 'r') as f:
                if filename.lower().endswith('.json'):
                    raw_data = json.load(f)
                elif filename.lower().endswith(('.yaml', '.yml')):
                    raw_data = yaml.safe_load(f)
                else:
                    # Try to detect format by content
                    content = f.read()
                    try:
                        raw_data = json.loads(content)
                    except json.JSONDecodeError:
                        try:
                            raw_data = yaml.safe_load(content)
                        except yaml.YAMLError:
                            raise ValueError("Unable to parse file as JSON or YAML")
            
            # Handle different data structures
            self.data = self._extract_resources(raw_data)
            
            # Reclassify VPC-related resources
            self._reclassify_vpc_resources()
            
            # Standardize service names and fix resource types
            self._standardize_service_names()
            self._fix_resource_types() 
            self._fix_id_and_name_parsing()
            self._fix_account_id_from_arn()
            self._deduplicate_resources()
            self._merge_cross_method_tags()
            
            # Enrich VPC/subnet information if enabled
            if self.enrich_vpc_info:
                self._enrich_vpc_subnet_info()
            
            # Collect all possible headers from the data
            for item in self.data:
                if isinstance(item, dict):
                    self._collect_headers(item)
            
            print(f"Loaded {len(self.data)} resources from {filename}")
            return self.data
            
        except FileNotFoundError:
            print(f"Error: File {filename} not found")
            sys.exit(1)
        except Exception as e:
            print(f"Error loading data: {e}")
            sys.exit(1)
    
    def _extract_resources(self, raw_data: Union[Dict, List]) -> List[Dict[str, Any]]:
        """Extract resources from different data structures."""
        resources = []
        
        # If it's already a list, assume it's a list of resources
        if isinstance(raw_data, list):
            return raw_data
        
        # If it's a dict, check for common resource array keys
        if isinstance(raw_data, dict):
            # Try to find resource arrays in common formats
            resource_keys = [
                'all_discovered_resources',  # Preferred - contains all resources
                'compliant_resources',
                'non_compliant_resources', 
                'untagged_resources',
                'resources',  # Generic key
                'items',      # Another common key
            ]
            
            # First try to get all resources from a single comprehensive array
            if 'all_discovered_resources' in raw_data and isinstance(raw_data['all_discovered_resources'], list):
                resources.extend(raw_data['all_discovered_resources'])
                print(f"Found {len(resources)} resources in 'all_discovered_resources'")
                return resources
            
            # Otherwise, collect from multiple arrays
            for key in resource_keys:
                if key in raw_data and isinstance(raw_data[key], list):
                    resources.extend(raw_data[key])
                    print(f"Found {len(raw_data[key])} resources in '{key}'")
            
            # If we found resources in arrays, return them
            if resources:
                return resources
            
            # If no resource arrays found, maybe the whole dict is a single resource
            # Check if it has typical resource fields
            if any(field in raw_data for field in ['service', 'type', 'id', 'arn', 'region']):
                return [raw_data]
        
        # Fallback: return empty list
        print("Warning: No resources found in the input data")
        return []
    
    def _reclassify_vpc_resources(self):
        """Reclassify VPC-related resources from EC2 to VPC service."""
        vpc_resource_types = {
            'VPC', 'Subnet', 'SecurityGroup', 'Route-Table', 'Network-Interface',
            'Vpc-Endpoint', 'Vpc-Flow-Log', 'Dhcp-Options', 'Transit-Gateway-Attachment',
            'Network-Insights-Path', 'Internet-Gateway', 'Nat-Gateway', 'Network-Acl'
        }
        
        reclassified_count = 0
        for resource in self.data:
            if isinstance(resource, dict) and resource.get('service') == 'EC2':
                resource_type = resource.get('type', '')
                if resource_type in vpc_resource_types:
                    resource['service'] = 'VPC'
                    reclassified_count += 1
        
        if reclassified_count > 0:
            print(f"Reclassified {reclassified_count} VPC-related resources from EC2 to VPC service")
    
    def _standardize_service_names(self):
        """Standardize service names to avoid duplicates."""
        standardizations = {
            'CloudFormation': 'CLOUDFORMATION',
            'Lambda': 'LAMBDA',
            # Add more as needed
        }
        
        standardized_count = 0
        for resource in self.data:
            if isinstance(resource, dict):
                service = resource.get('service', '')
                if service in standardizations:
                    resource['service'] = standardizations[service]
                    standardized_count += 1
        
        if standardized_count > 0:
            print(f"Standardized {standardized_count} service names")
    
    def _fix_resource_types(self):
        """Fix resource type inconsistencies."""
        fixed_count = 0
        
        for resource in self.data:
            if isinstance(resource, dict):
                service = resource.get('service', '')
                res_type = resource.get('type', '')
                arn = resource.get('arn', '')
                
                # Fix S3 bucket types (S3 bucket ARNs use arn:aws:s3:::bucket-name format)
                if service == 'S3' and arn.startswith('arn:aws:s3:::') and res_type != 'Bucket':
                    resource['type'] = 'Bucket'
                    fixed_count += 1
                
                # Fix SNS topic types (SNS topic ARNs use arn:aws:sns:region:account:topic-name format)
                elif service == 'SNS' and ':sns:' in arn and res_type != 'Topic':
                    resource['type'] = 'Topic'
                    fixed_count += 1
                
                # Fix SQS queue types (SQS queue ARNs use arn:aws:sqs:region:account:queue-name format)
                elif service == 'SQS' and ':sqs:' in arn and res_type != 'Queue':
                    resource['type'] = 'Queue'
                    fixed_count += 1
                
                # Fix CloudWatch alarm types (they all have generic "alarm" as ID)
                elif service == 'CLOUDWATCH' and res_type == 'Alarm' and ':alarm:' in arn:
                    # Extract alarm name from ARN and use it as a more unique identifier
                    alarm_name = arn.split(':alarm:')[-1] if ':alarm:' in arn else res_type
                    # Keep type as "Alarm" but update ID to be more specific
                    if resource.get('id') == 'alarm':
                        resource['id'] = alarm_name
        
        if fixed_count > 0:
            print(f"Fixed {fixed_count} resource type inconsistencies")
    
    def _fix_id_and_name_parsing(self):
        """Fix ID and name parsing issues by extracting correct values from ARNs."""
        fixed_count = 0
        
        for resource in self.data:
            if isinstance(resource, dict):
                arn = resource.get('arn', '')
                service = resource.get('service', '')
                current_id = resource.get('id', '')
                current_name = resource.get('name', '')
                
                if arn and arn.startswith('arn:aws:'):
                    correct_id, correct_name = self._extract_id_name_from_arn(arn, service)
                    
                    # Fix ID if current ID is generic or incorrect
                    if correct_id and (not current_id or current_id in [
                        'function', 'cluster', 'log-group', 'secret', 'parametergroup',
                        'subnetgroup', 'snapshot', 'alarm', 'topic', 'queue'
                    ] or current_id != correct_id):
                        resource['id'] = correct_id
                        fixed_count += 1
                    
                    # Fix name if missing or incorrect
                    if correct_name and (not current_name or current_name != correct_name):
                        resource['name'] = correct_name
                        fixed_count += 1
        
        if fixed_count > 0:
            print(f"Fixed {fixed_count} ID and name parsing issues")
    
    def _extract_id_name_from_arn(self, arn: str, service: str) -> tuple:
        """Extract correct ID and name from ARN based on service type."""
        if not arn or not arn.startswith('arn:aws:'):
            return None, None
        
        try:
            # Standard ARN format: arn:aws:service:region:account:resource
            parts = arn.split(':')
            if len(parts) < 6:
                return None, None
            
            # Service-specific parsing logic
            if service == 'S3':
                # arn:aws:s3:::bucket-name
                if len(parts) >= 6:
                    bucket_name = parts[5]
                    return bucket_name, bucket_name
            
            elif service == 'LAMBDA':
                # arn:aws:lambda:region:account:function:function-name
                if len(parts) >= 7 and parts[5] == 'function':
                    func_name = parts[6]
                    return func_name, func_name
            
            elif service == 'ELASTICACHE':
                # arn:aws:elasticache:region:account:cluster:cluster-name
                # arn:aws:elasticache:region:account:parametergroup:group-name
                if len(parts) >= 7:
                    resource_name = parts[6]
                    return resource_name, resource_name
            
            elif service == 'LOGS':
                # arn:aws:logs:region:account:log-group:/log/group/name
                if len(parts) >= 7 and parts[5] == 'log-group':
                    log_group_name = parts[6]
                    return log_group_name, log_group_name
            
            elif service == 'SECRETSMANAGER':
                # arn:aws:secretsmanager:region:account:secret:secret-name-random
                if len(parts) >= 7 and parts[5] == 'secret':
                    secret_full = parts[6]
                    # Remove the random suffix (last 6 characters after last dash)
                    if '-' in secret_full and len(secret_full.split('-')[-1]) == 6:
                        secret_name = '-'.join(secret_full.split('-')[:-1])
                    else:
                        secret_name = secret_full
                    return secret_name, secret_name
            
            elif service == 'SNS':
                # arn:aws:sns:region:account:topic-name
                if len(parts) >= 6:
                    topic_name = parts[5]
                    return topic_name, topic_name
            
            elif service == 'SQS':
                # arn:aws:sqs:region:account:queue-name
                if len(parts) >= 6:
                    queue_name = parts[5]
                    return queue_name, queue_name
            
            elif service == 'IAM':
                # arn:aws:iam::account:policy/policy-name
                # arn:aws:iam::account:role/role-name
                # arn:aws:iam::account:user/user-name
                if len(parts) >= 6:
                    resource_part = parts[5]
                    if '/' in resource_part:
                        resource_name = resource_part.split('/')[-1]
                        return resource_name, resource_name
                    return resource_part, resource_part
            
            elif service == 'GLUE':
                # arn:aws:glue:region:account:crawler/crawler-name
                # arn:aws:glue:region:account:job/job-name
                if len(parts) >= 6:
                    resource_part = parts[5] 
                    if '/' in resource_part:
                        resource_name = resource_part.split('/')[-1]
                        return resource_name, resource_name
                    return resource_part, resource_part
            
            elif service == 'SSM':
                # arn:aws:ssm:region:account:session/session-id
                # arn:aws:ssm:region:account:parameter/parameter-name
                if len(parts) >= 6:
                    resource_part = parts[5]
                    if '/' in resource_part:
                        resource_name = resource_part.split('/', 1)[1]  # Keep full path after first /
                        return resource_name, resource_name
                    return resource_part, resource_part
            
            elif service == 'CLOUDWATCH':
                # arn:aws:cloudwatch:region:account:alarm:alarm-name
                if len(parts) >= 7 and parts[5] == 'alarm':
                    alarm_name = parts[6]
                    return alarm_name, alarm_name
            
            elif service in ['VPC', 'EC2']:
                # arn:aws:ec2:region:account:vpc/vpc-id
                # arn:aws:ec2:region:account:subnet/subnet-id
                # arn:aws:ec2:region:account:security-group/sg-id
                if len(parts) >= 6:
                    resource_part = parts[5]
                    if '/' in resource_part:
                        resource_name = resource_part.split('/')[-1]
                        return resource_name, resource_name
                    return resource_part, resource_part
            
            # Generic fallback
            if len(parts) >= 6:
                resource_part = parts[5]
                if ':' in resource_part and len(parts) >= 7:
                    # Format like resource-type:resource-name (use parts[6])
                    resource_name = parts[6]
                    return resource_name, resource_name
                elif '/' in resource_part:
                    # Format like resource-type/resource-name
                    resource_name = resource_part.split('/')[-1]
                    return resource_name, resource_name
                else:
                    return resource_part, resource_part
                    
        except Exception:
            # If parsing fails, return original values
            return None, None
    
    def _fix_account_id_from_arn(self):
        """Extract and set account_id from ARN for resources missing this field."""
        fixed_count = 0
        
        for resource in self.data:
            if isinstance(resource, dict):
                current_account_id = resource.get('account_id')
                arn = resource.get('arn', '')
                
                # Fix missing account_id
                if not current_account_id and arn and arn.startswith('arn:aws:'):
                    extracted_account_id = self._extract_account_id_from_arn(arn)
                    
                    if extracted_account_id:
                        resource['account_id'] = extracted_account_id
                        fixed_count += 1
        
        if fixed_count > 0:
            print(f"Fixed {fixed_count} missing account_id fields from ARNs")
    
    def _extract_account_id_from_arn(self, arn: str) -> str:
        """Extract account ID from ARN."""
        if not arn or not arn.startswith('arn:aws:'):
            return None
        
        try:
            # Standard ARN format: arn:aws:service:region:account-id:resource
            parts = arn.split(':')
            
            # Handle special cases
            service = parts[2] if len(parts) > 2 else ''
            
            if service == 's3':
                # S3 ARNs don't contain account ID directly
                # arn:aws:s3:::bucket-name
                # We need to infer it from other resources or leave it empty
                # For now, we'll try to find it from other resources in the same dataset
                return self._infer_account_id_for_s3()
            
            elif service == 'iam':
                # IAM ARNs: arn:aws:iam::account-id:resource
                # Note the empty region field (double colon)
                if len(parts) >= 5:
                    return parts[4] if parts[4] else None
            
            else:
                # Standard format: arn:aws:service:region:account-id:resource
                if len(parts) >= 5:
                    return parts[4] if parts[4] else None
                    
        except Exception:
            return None
        
        return None
    
    def _infer_account_id_for_s3(self) -> str:
        """Infer account ID for S3 resources from other resources in the dataset."""
        # Look for any resource that has an account_id already set
        for resource in self.data:
            if isinstance(resource, dict):
                account_id = resource.get('account_id')
                if account_id and account_id.isdigit() and len(account_id) == 12:
                    return account_id
        
        # Fallback: extract from any non-S3 ARN in the dataset
        for resource in self.data:
            if isinstance(resource, dict):
                arn = resource.get('arn', '')
                service = resource.get('service', '')
                if arn and service != 'S3':
                    extracted = self._extract_account_id_from_arn(arn)
                    if extracted and extracted.isdigit() and len(extracted) == 12:
                        return extracted
        
        return None
    
    def _deduplicate_resources(self):
        """Remove duplicate resources keeping the one with more complete information."""
        seen_resources = {}
        deduplicated = []
        duplicate_count = 0
        
        for resource in self.data:
            if isinstance(resource, dict):
                # Create a unique key based on ARN (most reliable) or fallback to service+id+region
                arn = resource.get('arn', '')
                if arn:
                    key = arn
                else:
                    service = resource.get('service', '')
                    res_id = resource.get('id', '')
                    region = resource.get('region', '')
                    key = f"{service}:{res_id}:{region}"
                
                if key in seen_resources:
                    # We have a duplicate - compare and keep the better one
                    existing = seen_resources[key]
                    current = resource
                    
                    # Prefer resources with more complete information
                    existing_score = self._score_resource_completeness(existing)
                    current_score = self._score_resource_completeness(current)
                    
                    if current_score > existing_score:
                        # Replace existing with current
                        seen_resources[key] = current
                        # Remove existing from deduplicated list and add current
                        deduplicated = [r for r in deduplicated if r != existing]
                        deduplicated.append(current)
                    
                    duplicate_count += 1
                else:
                    seen_resources[key] = resource
                    deduplicated.append(resource)
        
        if duplicate_count > 0:
            print(f"Removed {duplicate_count} duplicate resources")
            print(f"Kept {len(deduplicated)} unique resources")
        
        self.data = deduplicated
    
    def _score_resource_completeness(self, resource: Dict[str, Any]) -> int:
        """Score a resource based on completeness of information."""
        score = 0
        
        # Heavily prioritize resources with tags since they're crucial for BOM analysis
        tags = resource.get('tags', {}) or {}
        tag_count = len(tags)
        if tag_count > 0:
            score += 20 + tag_count  # Base 20 points plus 1 per tag
        
        # Prefer resources discovered via certain methods, but tags override this
        discovered_via = resource.get('discovered_via', '')
        if discovered_via == 'ResourceGroupsTaggingAPI':
            score += 10  # Tagging API is most likely to have tag information
        elif discovered_via == 'AWSConfig':
            score += 8   # Config has good config data but often missing tags
        elif discovered_via == 'ServiceSpecificAPI':
            score += 6   # Service APIs are comprehensive but variable tag coverage
        
        # Score based on available fields
        important_fields = ['name', 'region', 'account_id']
        for field in important_fields:
            if resource.get(field):
                score += 1
        
        return score
    
    def _merge_cross_method_tags(self):
        """Attempt to enrich AWSConfig resources with tag patterns from ResourceGroupsTaggingAPI resources."""
        # Create tag pattern analysis from tagged resources
        tag_patterns = self._analyze_tag_patterns()
        
        # Apply patterns to untagged AWSConfig resources
        enriched_count = self._apply_tag_patterns(tag_patterns)
        
        if enriched_count > 0:
            print(f"Enhanced {enriched_count} AWSConfig resources with inferred tag information")
    
    def _analyze_tag_patterns(self):
        """Analyze common tag patterns from tagged resources."""
        patterns = {
            'service_region': {},  # Tags common to service+region combinations
            'service_type': {},    # Tags common to service+type combinations
            'account_wide': {}     # Tags that appear frequently across the account
        }
        
        tagged_resources = [r for r in self.data if len(r.get('tags', {}) or {}) > 0]
        
        for resource in tagged_resources:
            service = resource.get('service', '')
            region = resource.get('region', '')
            res_type = resource.get('type', '')
            tags = resource.get('tags', {}) or {}
            
            # Track patterns by service+region
            sr_key = f"{service}:{region}"
            if sr_key not in patterns['service_region']:
                patterns['service_region'][sr_key] = {}
            for tag_key, tag_value in tags.items():
                if tag_key not in patterns['service_region'][sr_key]:
                    patterns['service_region'][sr_key][tag_key] = {}
                if tag_value not in patterns['service_region'][sr_key][tag_key]:
                    patterns['service_region'][sr_key][tag_key][tag_value] = 0
                patterns['service_region'][sr_key][tag_key][tag_value] += 1
            
            # Track patterns by service+type
            st_key = f"{service}:{res_type}"
            if st_key not in patterns['service_type']:
                patterns['service_type'][st_key] = {}
            for tag_key, tag_value in tags.items():
                if tag_key not in patterns['service_type'][st_key]:
                    patterns['service_type'][st_key][tag_key] = {}
                if tag_value not in patterns['service_type'][st_key][tag_key]:
                    patterns['service_type'][st_key][tag_key][tag_value] = 0
                patterns['service_type'][st_key][tag_key][tag_value] += 1
            
            # Track account-wide patterns
            for tag_key, tag_value in tags.items():
                if tag_key not in patterns['account_wide']:
                    patterns['account_wide'][tag_key] = {}
                if tag_value not in patterns['account_wide'][tag_key]:
                    patterns['account_wide'][tag_key][tag_value] = 0
                patterns['account_wide'][tag_key][tag_value] += 1
        
        return patterns
    
    def _apply_tag_patterns(self, patterns):
        """Apply tag patterns to untagged AWSConfig resources."""
        enriched_count = 0
        
        for resource in self.data:
            if (isinstance(resource, dict) and 
                resource.get('discovered_via') == 'AWSConfig' and 
                len(resource.get('tags', {}) or {}) == 0):
                
                service = resource.get('service', '')
                region = resource.get('region', '')
                res_type = resource.get('type', '')
                inferred_tags = {}
                
                # Try to infer tags from service+region patterns
                sr_key = f"{service}:{region}"
                if sr_key in patterns['service_region']:
                    for tag_key, tag_values in patterns['service_region'][sr_key].items():
                        # Use the most common tag value for this service+region
                        most_common_value = max(tag_values.items(), key=lambda x: x[1])
                        if most_common_value[1] >= 2:  # Appears at least twice
                            inferred_tags[tag_key] = most_common_value[0]
                
                # Try to infer tags from service+type patterns
                st_key = f"{service}:{res_type}"
                if st_key in patterns['service_type']:
                    for tag_key, tag_values in patterns['service_type'][st_key].items():
                        if tag_key not in inferred_tags:  # Don't override service+region tags
                            most_common_value = max(tag_values.items(), key=lambda x: x[1])
                            if most_common_value[1] >= 2:  # Appears at least twice
                                inferred_tags[tag_key] = most_common_value[0]
                
                # Add highly common account-wide tags
                for tag_key, tag_values in patterns['account_wide'].items():
                    if tag_key not in inferred_tags:
                        most_common_value = max(tag_values.items(), key=lambda x: x[1])
                        total_tagged_resources = len([r for r in self.data if len(r.get('tags', {}) or {}) > 0])
                        # Only add if this tag appears in >50% of tagged resources
                        if most_common_value[1] > total_tagged_resources * 0.5:
                            inferred_tags[tag_key] = most_common_value[0]
                
                # Apply inferred tags
                if inferred_tags:
                    # Mark them as inferred to distinguish from real tags
                    for tag_key in list(inferred_tags.keys()):
                        inferred_tags[f"{tag_key}"] = inferred_tags[tag_key]
                    
                    resource['tags'] = inferred_tags
                    resource['tags_inferred'] = True  # Flag to indicate these are inferred
                    enriched_count += 1
        
        return enriched_count
    
    def _collect_headers(self, item: Dict[str, Any], prefix: str = ""):
        """Recursively collect all possible headers from nested dictionaries."""
        for key, value in item.items():
            header = f"{prefix}{key}" if prefix else key
            
            if isinstance(value, dict):
                # For nested dictionaries, create flattened headers
                self._collect_headers(value, f"{header}.")
            elif isinstance(value, list):
                # For lists, just add the header (we'll stringify the list)
                self.headers.add(header)
            else:
                self.headers.add(header)
    
    def _flatten_dict(self, item: Dict[str, Any], prefix: str = "") -> Dict[str, str]:
        """Flatten nested dictionaries for tabular output."""
        flattened = {}
        
        for key, value in item.items():
            header = f"{prefix}{key}" if prefix else key
            
            if isinstance(value, dict):
                flattened.update(self._flatten_dict(value, f"{header}."))
            elif isinstance(value, list):
                # Convert lists to comma-separated strings
                flattened[header] = ", ".join(str(v) for v in value)
            elif value is None:
                flattened[header] = ""
            else:
                flattened[header] = str(value)
        
        return flattened
    
    def _enrich_vpc_subnet_info(self):
        """Enrich resources with VPC and subnet names."""
        if not self.session:
            return
        
        print("Enriching VPC and subnet information...")
        
        for item in self.data:
            if isinstance(item, dict):
                region = item.get('region', '')
                vpc_id = item.get('vpc_id', '')
                subnet_id = item.get('subnet_id', '')
                
                # Skip global resources or resources without VPC info
                if region in ['global', ''] or (not vpc_id and not subnet_id):
                    continue
                
                try:
                    # Get VPC name if vpc_id exists
                    if vpc_id and vpc_id != '':
                        vpc_name = self._get_vpc_name(region, vpc_id)
                        if vpc_name:
                            item['vpc_name'] = vpc_name
                    
                    # Get subnet name if subnet_id exists
                    if subnet_id and subnet_id != '':
                        subnet_name = self._get_subnet_name(region, subnet_id)
                        if subnet_name:
                            item['subnet_name'] = subnet_name
                
                except Exception as e:
                    # Silently continue if we can't get VPC/subnet info
                    continue
    
    def _get_vpc_name(self, region: str, vpc_id: str) -> Optional[str]:
        """Get VPC name from tags."""
        cache_key = f"{region}:{vpc_id}"
        
        if cache_key in self.vpc_cache:
            return self.vpc_cache[cache_key]
        
        try:
            ec2 = self.session.client('ec2', region_name=region)
            response = ec2.describe_vpcs(VpcIds=[vpc_id])
            
            if response['Vpcs']:
                vpc = response['Vpcs'][0]
                # Look for Name tag
                for tag in vpc.get('Tags', []):
                    if tag['Key'] == 'Name':
                        name = tag['Value']
                        self.vpc_cache[cache_key] = name
                        return name
                
                # If no Name tag, cache empty result
                self.vpc_cache[cache_key] = None
                return None
        
        except ClientError:
            # Cache None result for failed lookups to avoid repeated attempts
            self.vpc_cache[cache_key] = None
            return None
    
    def _get_subnet_name(self, region: str, subnet_id: str) -> Optional[str]:
        """Get subnet name from tags."""
        cache_key = f"{region}:{subnet_id}"
        
        if cache_key in self.vpc_cache:
            return self.vpc_cache[cache_key]
        
        try:
            ec2 = self.session.client('ec2', region_name=region)
            response = ec2.describe_subnets(SubnetIds=[subnet_id])
            
            if response['Subnets']:
                subnet = response['Subnets'][0]
                # Look for Name tag
                for tag in subnet.get('Tags', []):
                    if tag['Key'] == 'Name':
                        name = tag['Value']
                        self.vpc_cache[cache_key] = name
                        return name
                
                # If no Name tag, cache empty result
                self.vpc_cache[cache_key] = None
                return None
        
        except ClientError:
            # Cache None result for failed lookups to avoid repeated attempts
            self.vpc_cache[cache_key] = None
            return None
    
    def to_csv(self, output_filename: str):
        """Convert data to CSV format with AWS service as a column."""
        try:
            # Sort headers for consistent output, ensuring 'service' comes first
            sorted_headers = sorted(list(self.headers))
            if 'service' in sorted_headers:
                sorted_headers.remove('service')
                sorted_headers.insert(0, 'service')
            
            with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=sorted_headers)
                writer.writeheader()
                
                for item in self.data:
                    if isinstance(item, dict):
                        flattened = self._flatten_dict(item)
                        # Ensure all headers are present (fill missing with empty string)
                        row = {header: flattened.get(header, '') for header in sorted_headers}
                        writer.writerow(row)
            
            print(f"CSV export saved to {output_filename} (AWS service included as column)")
            
        except Exception as e:
            print(f"Error creating CSV: {e}")
            sys.exit(1)
    
    def to_excel(self, output_filename: str):
        """Convert data to Excel format with separate sheets per AWS service."""
        try:
            # Try to import openpyxl, fall back to CSV if not available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
            except ImportError:
                print("Warning: openpyxl not available. Falling back to CSV format.")
                csv_filename = output_filename.replace('.xlsx', '.csv')
                self.to_csv(csv_filename)
                return
            
            # Group resources by service
            resources_by_service = {}
            for item in self.data:
                if isinstance(item, dict):
                    service = item.get('service', 'Unknown')
                    if service not in resources_by_service:
                        resources_by_service[service] = []
                    resources_by_service[service].append(item)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet first
            self._create_summary_sheet(wb, resources_by_service)
            
            # Create a sheet for each service
            for service_name in sorted(resources_by_service.keys()):
                service_resources = resources_by_service[service_name]
                self._create_service_sheet(wb, service_name, service_resources)
            
            # Save workbook
            wb.save(output_filename)
            print(f"Excel export saved to {output_filename}")
            print(f"Created {len(resources_by_service)} service sheets + summary sheet")
            
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            sys.exit(1)
    
    def _create_summary_sheet(self, wb: 'Workbook', resources_by_service: Dict[str, List[Dict]]):
        """Create summary sheet with overview statistics."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        summary_ws = wb.create_sheet("Summary", 0)
        
        # Title
        summary_ws.append(["AWS Resource Inventory Summary"])
        summary_ws.cell(1, 1).font = Font(bold=True, size=16)
        summary_ws.append([])
        
        # Basic info
        summary_ws.append(["Generated on:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_ws.append(["Total Resources:", len(self.data)])
        summary_ws.append(["Total Services:", len(resources_by_service)])
        summary_ws.append([])
        
        # Service breakdown table
        summary_ws.append(["Service Breakdown:"])
        summary_ws.cell(7, 1).font = Font(bold=True, size=12)
        summary_ws.append([])
        
        # Table headers
        headers = ["Service", "Resource Count", "Resource Types"]
        summary_ws.append(headers)
        
        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = summary_ws.cell(9, col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Service data
        for service_name in sorted(resources_by_service.keys()):
            service_resources = resources_by_service[service_name]
            resource_types = set(item.get('type', 'Unknown') for item in service_resources)
            
            summary_ws.append([
                service_name,
                len(service_resources),
                ", ".join(sorted(resource_types))
            ])
        
        # Auto-adjust column widths
        for col_num in range(1, 4):
            column_letter = get_column_letter(col_num)
            max_length = 10
            
            for row in summary_ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            summary_ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
    
    def _create_service_sheet(self, wb: 'Workbook', service_name: str, service_resources: List[Dict]):
        """Create a sheet for a specific AWS service."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Sanitize sheet name (Excel sheet names have restrictions)
        safe_service_name = service_name.replace('/', '_').replace('\\', '_')[:31]
        ws = wb.create_sheet(safe_service_name)
        
        # Collect headers specific to this service
        service_headers = set()
        for item in service_resources:
            if isinstance(item, dict):
                self._collect_headers_for_service(item, service_headers)
        
        # Sort headers, prioritizing important ones first
        priority_headers = ['service', 'type', 'region', 'id', 'name', 'arn', 'vpc_id', 'vpc_name', 'subnet_id', 'subnet_name']
        sorted_headers = []
        
        # Add priority headers first (if they exist)
        for header in priority_headers:
            if header in service_headers:
                sorted_headers.append(header)
                service_headers.remove(header)
        
        # Add remaining headers
        sorted_headers.extend(sorted(service_headers))
        
        # Write headers with formatting
        for col_num, header in enumerate(sorted_headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write data rows
        for row_num, item in enumerate(service_resources, 2):
            if isinstance(item, dict):
                flattened = self._flatten_dict(item)
                for col_num, header in enumerate(sorted_headers, 1):
                    value = flattened.get(header, '')
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col_num, header in enumerate(sorted_headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(header)
            
            # Check data in this column to find max length
            for row_num in range(2, min(102, len(service_resources) + 2)):  # Check first 100 rows
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width (max 50 to avoid extremely wide columns)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _collect_headers_for_service(self, item: Dict[str, Any], headers_set: set, prefix: str = ""):
        """Collect headers into a specific set (for service-specific sheets)."""
        for key, value in item.items():
            header = f"{prefix}{key}" if prefix else key
            
            if isinstance(value, dict):
                # For nested dictionaries, create flattened headers
                self._collect_headers_for_service(value, headers_set, f"{header}.")
            elif isinstance(value, list):
                # For lists, just add the header (we'll stringify the list)
                headers_set.add(header)
            else:
                headers_set.add(header)


def main():
    parser = argparse.ArgumentParser(
        description='AWS InvenTag BOM Converter - Professional AWS resource inventory to Excel/CSV converter',
        epilog='Part of AWS InvenTag: Python tool for AWS cloud inventory and tagging compliance.'
    )
    parser.add_argument('--input', '-i', required=True, 
                       help='Input JSON or YAML file')
    parser.add_argument('--output', '-o', required=True, 
                       help='Output filename')
    parser.add_argument('--format', choices=['excel', 'csv'], default='excel',
                       help='Output format (default: excel)')
    parser.add_argument('--no-vpc-enrichment', action='store_true',
                       help='Skip VPC/subnet name enrichment (faster but less detailed)')
    
    args = parser.parse_args()
    
    # Validate input file exists
    if not os.path.exists(args.input):
        print(f"Error: Input file {args.input} does not exist")
        sys.exit(1)
    
    # Initialize converter
    enrich_vpc = not args.no_vpc_enrichment
    converter = BOMConverter(enrich_vpc_info=enrich_vpc)
    
    if enrich_vpc:
        print("VPC/subnet name enrichment enabled (use --no-vpc-enrichment to disable)")
    else:
        print("VPC/subnet name enrichment disabled")
    
    # Load data
    converter.load_data(args.input)
    
    # Convert based on format
    if args.format == 'excel':
        # Ensure output has .xlsx extension
        if not args.output.lower().endswith('.xlsx'):
            args.output += '.xlsx'
        converter.to_excel(args.output)
    elif args.format == 'csv':
        # Ensure output has .csv extension
        if not args.output.lower().endswith('.csv'):
            args.output += '.csv'
        converter.to_csv(args.output)


if __name__ == "__main__":
    main()